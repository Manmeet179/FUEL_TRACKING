import streamlit as st
import pandas as pd
from datetime import datetime
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import bcrypt
from dotenv import load_dotenv
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet

st.set_page_config(
    page_title="Petrol Expense Tracker",  # Change tab name
    page_icon="fimg_2.ico"               # Path to your icon (favicon)
)

# ====== LOAD ENV ======
load_dotenv()

# ====== CONFIG ======
allowed_users = {
    os.getenv("USER1_NAME"): {
        "name": os.getenv("USER1_EMAIL"),
        "password_hash": os.getenv("USER1_HASH")
    },
    os.getenv("USER2_EMAIL"): {
        "name": os.getenv("USER2_NAME"),
        "password_hash": os.getenv("USER2_HASH")
    }
}

home_office_km_dict = {
    os.getenv("USER1_NAME"): 8,
    os.getenv("USER2_NAME"): 30
}

# ====== SESSION INIT ======
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.username = None

if "show_edit" not in st.session_state:
    st.session_state.show_edit = False
if "show_delete_confirm" not in st.session_state:
    st.session_state.show_delete_confirm = False
if "edit_index" not in st.session_state:
    st.session_state.edit_index = None
if "delete_index" not in st.session_state:
    st.session_state.delete_index = None

# ====== FILE HANDLING ======
def get_file_path(name):
    folder = "petrol_expense_files"
    os.makedirs(folder, exist_ok=True)
    month_str = datetime.today().strftime("%b")  # Aug, Sep, etc.
    filename = f"PE-{month_str}-{name}.xlsx"
    return os.path.join(folder, filename)

def load_data(file_path):
    if os.path.exists(file_path):
        try:
            return pd.read_excel(file_path)
        except:
            return pd.DataFrame(columns=["Sr", "Date", "Particulars - Travelling Details", "Purpose", "KMS", "Total INR"])
    else:
        return pd.DataFrame(columns=["Sr", "Date", "Particulars - Travelling Details", "Purpose", "KMS", "Total INR"])

def save_excel_formatted(df, name, total_km_sum, total_inr_sum):
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Petrol Expense"

    ws.merge_cells("B1:F1")
    ws["B1"] = f"Petrol Expense Summary - {datetime.today().strftime('%b-%Y')}"
    ws["B1"].font = Font(bold=True, size=14)
    ws["B1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("B2:E2")
    ws["B2"] = f"Employee Name: {name}"
    ws["B2"].font = Font(bold=True)

    ws["F2"] = f"Date: {datetime.today().strftime('%d.%m.%y')}"
    ws["F2"].alignment = Alignment(horizontal="right")

    ws.merge_cells("B3:F3")
    ws["B3"] = "Petrol Conveyance: 4 INR / Kms for 2-Wheeler"

    columns = list(df.columns)
    for col_num, col_name in enumerate(columns, 1):
        cell = ws.cell(row=4, column=col_num, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

    for row_num, row in enumerate(df.values.tolist(), 5):
        for col_num, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))

    total_row = len(df) + 5
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(row=total_row, column=1, value="Total").alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=5, value=total_km_sum)
    ws.cell(row=total_row, column=6, value=total_inr_sum)

    for col in range(1, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

    for i in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_pdf_from_df(df, name, total_km_sum, total_inr_sum):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"<b>Petrol Expense Summary - {datetime.today().strftime('%b %Y')}</b>", styles["Title"]))
    elements.append(Paragraph(f"<b>Employee Name:</b> {name}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Date:</b> {datetime.today().strftime('%d-%m-%Y')}", styles["Normal"]))
    elements.append(Paragraph("Petrol Conveyance: ₹4 / KM (2-Wheeler)", styles["Normal"]))
    elements.append(Spacer(1, 12))

    table_data = [list(df.columns)]
    for row in df.values.tolist():
        table_data.append(row)
    table_data.append(["", "", "", "Total", total_km_sum, total_inr_sum])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 24))
    elements.append(Paragraph("Approved by: ____________________", styles["Normal"]))
    doc.build(elements)

    buffer.seek(0)
    return buffer

# ====== LOGIN SECTION ======
def login_section():
    st.title("🔒 Petrol Expense Login")
    email = st.text_input("Email")
    password = st.text_input("Password", type="password")

    if st.button("Login"):
        user = allowed_users.get(email)
        if user and user["password_hash"] and bcrypt.checkpw(password.encode(), user["password_hash"].encode()):
            st.session_state.logged_in = True
            st.session_state.username = user["name"]
            st.rerun()
        else:
            st.error("Invalid email or password ❌")

# ====== MAIN APP ======
def main_app():
    name = st.session_state.username
    file_path = get_file_path(name)

    # Load data from file on first load or when month changes
    if "df_data" not in st.session_state or st.session_state.get("loaded_file") != file_path:
        st.session_state.df_data = load_data(file_path)
        st.session_state.loaded_file = file_path

    df = st.session_state.df_data

    st.title("Petrol Expense Entry")

    col1, col2 = st.columns([1,1])
    with col1:
        st.markdown(f"📅 {datetime.today().strftime('%A, %d %B %Y')}")
    with col2:
        st.markdown(f"<div style='text-align:right;'>👤 {name}</div>", unsafe_allow_html=True)

    total_km = st.number_input("Enter Total KM (Today's travel)", min_value=0.0, step=0.1)
    default_home_km = home_office_km_dict.get(name, 0)
    home_office_km = st.number_input("Enter Home to Office KM", min_value=0.0, step=0.1, value=float(default_home_km))
    particulars = st.text_input("Particulars - Travelling Details")
    purpose = st.text_input("Purpose")

    if st.button("Save Entry"):
        if not particulars.strip():
            st.error("⚠ Please enter the travelling details.")
        elif not purpose.strip():
            st.error("⚠ Please enter the purpose.")
        elif total_km < 1 or total_km > 999:
            st.error("⚠ Please enter Valid Total KM")
        else:
            date_today = datetime.today().strftime("%d-%b")
            travel_km = max(total_km - home_office_km, 0)
            total_inr = travel_km * 4

            new_row = {
                "Sr": len(df) + 1,
                "Date": date_today,
                "Particulars - Travelling Details": particulars,
                "Purpose": purpose,
                "KMS": travel_km,
                "Total INR": total_inr
            }

            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            df["Sr"] = range(1, len(df) + 1)
            df.to_excel(file_path, index=False)
            st.session_state.df_data = df
            st.success(f"✅ Entry Saved: {travel_km} KM → ₹{total_inr}")

    st.subheader(f"Summary for {name} - {datetime.today().strftime('%b %Y')}")
    st.write("Petrol Conveyance: ₹4 / KM (2-Wheeler)")

    df = st.session_state.df_data

    # Display table with Edit/Delete buttons vertically aligned per row
    for idx, row in df.iterrows():
        st.markdown(f"**Sr:** {row['Sr']}")
        st.markdown(f"**Date:** {row['Date']}")
        st.markdown(f"**Particulars - Travelling Details:** {row['Particulars - Travelling Details']}")
        st.markdown(f"**Purpose:** {row['Purpose']}")
        st.markdown(f"**KMS:** {row['KMS']}")
        st.markdown(f"**Total INR:** ₹{row['Total INR']}")

        cols = st.columns([1,1])
        if cols[0].button("✏️ Edit", key=f"edit_{idx}"):
            st.session_state.edit_index = idx
            st.session_state.show_edit = True
            st.experimental_rerun()
        if cols[1].button("🗑️ Delete", key=f"delete_{idx}"):
            st.session_state.delete_index = idx
            st.session_state.show_delete_confirm = True
            st.rerun()

        st.markdown("---")

    # Edit form
    if st.session_state.show_edit:
        idx = st.session_state.edit_index
        row = df.loc[idx]

        st.write(f"Editing Sr: {row['Sr']}")

        with st.form("edit_form_multi"):
            col_date, col_total_km, col_home_km = st.columns(3)
            new_date = col_date.text_input("Date (e.g. 15-Aug)", value=row["Date"])
            old_home_km = home_office_km_dict.get(name, 0)
            new_total_km = col_total_km.number_input(
                "Total KM (Today's travel)",
                min_value=0.0,
                step=0.1,
                value=float(row["KMS"] + old_home_km)
            )
            new_home_km = col_home_km.number_input(
                "Home to Office KM",
                min_value=0.0,
                step=0.1,
                value=float(old_home_km)
            )

            particulars_input = st.text_input("Particulars - Travelling Details", value=row["Particulars - Travelling Details"])
            purpose_input = st.text_input("Purpose", value=row["Purpose"])

            submitted = st.form_submit_button("Save")
            cancelled = st.form_submit_button("Cancel")

            if submitted:
                travel_km = max(new_total_km - new_home_km, 0)
                total_inr = travel_km * 4

                df.at[idx, "Date"] = new_date
                df.at[idx, "Particulars - Travelling Details"] = particulars_input
                df.at[idx, "Purpose"] = purpose_input
                df.at[idx, "KMS"] = travel_km
                df.at[idx, "Total INR"] = total_inr

                df.to_excel(file_path, index=False)
                st.session_state.df_data = df
                st.success("✅ Entry updated!")
                st.session_state.show_edit = False
                st.rerun()

            elif cancelled:
                st.session_state.show_edit = False
                st.rerun()

    # Delete confirmation popup
    if st.session_state.show_delete_confirm:
        idx = st.session_state.delete_index
        st.warning(f"Are you sure you want to delete Sr: {df.at[idx, 'Sr']} - Particulars: {df.at[idx, 'Particulars - Travelling Details']}?")
        col_yes, col_no = st.columns(2)
        if col_yes.button("Yes, Delete"):
            df = df.drop(idx).reset_index(drop=True)
            df["Sr"] = range(1, len(df) + 1)
            df.to_excel(file_path, index=False)
            st.session_state.df_data = df
            st.session_state.show_delete_confirm = False
            st.success("✅ Entry deleted!")
            st.experimental_rerun()
        if col_no.button("No, Cancel"):
            st.session_state.show_delete_confirm = False
            st.experimental_rerun()

    total_km_sum = df["KMS"].sum()
    total_inr_sum = df["Total INR"].sum()

    st.write(f"Total KMS: {total_km_sum}")
    st.write(f"Total INR: ₹{total_inr_sum}")

    excel_buffer = save_excel_formatted(df, name, total_km_sum, total_inr_sum)
    st.download_button(
        "⬇ Download Excel",
        data=excel_buffer,
        file_name=os.path.basename(file_path),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    pdf_buffer = generate_pdf_from_df(df, name, total_km_sum, total_inr_sum)
    st.download_button(
        "⬇ Download PDF",
        data=pdf_buffer,
        file_name=f"PE-{datetime.today().strftime('%b')}-{name}.pdf",
        mime="application/pdf"
    )

# ====== SIDEBAR LOGOUT ======
if st.session_state.logged_in:
    if st.sidebar.button("🚪 Logout"):
        st.session_state.logged_in = False
        st.session_state.username = None
        st.session_state.df_data = pd.DataFrame(
            columns=["Sr", "Date", "Particulars - Travelling Details", "Purpose", "KMS", "Total INR"]
        )
        st.rerun()

# ====== APP ENTRY POINT ======
if not st.session_state.logged_in:
    login_section()
else:
    main_app()

